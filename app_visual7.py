import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from pathlib import Path
import threading

import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# ---------- Utilidades ----------
MESES_PT_EN = {
    'jan': 'Jan', 'fev': 'Feb', 'mar': 'Mar', 'abr': 'Apr',
    'mai': 'May', 'jun': 'Jun', 'jul': 'Jul', 'ago': 'Aug',
    'set': 'Sep', 'out': 'Oct', 'nov': 'Nov', 'dez': 'Dec'
}

def converter_data_ptbr(mes_ano: str):
    if isinstance(mes_ao := mes_ano, str) and '/' in mes_ao:
        mes, ano = mes_ao.split('/')
        mes_en = MESES_PT_EN.get(mes.lower(), mes)
        return datetime.strptime(f"{mes_en}/{ano}", "%b/%y").replace(day=1)
    return pd.NaT

def corrigir_planilha_entrada(caminho_xlsx: Path) -> Path:
    """
    1) Renomeia aba ativa para 'Aba1'
    2) Se houver merge que intersecte A1:C2, desfaz, move valor A1 -> C1 e limpa A1
    3) Salva uma cópia temporária corrigida ao lado do original
    """
    wb = load_workbook(caminho_xlsx)
    ws = wb.active
    ws.title = "Aba1"

    try:
        ws.delete_cols(2, 1)  # apaga a coluna B
    except Exception:
        pass

    # Trata merges que atinjam A1:C2
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= 2 and rng.min_col <= 3 and rng.max_row >= 1 and rng.max_col >= 1:
            valor = ws["A1"].value
            ws.unmerge_cells(str(rng))
            ws["C1"] = valor
            ws["A1"] = None
            break

    temp_path = caminho_xlsx.with_name(f"{caminho_xlsx.stem}__corrigido.xlsx")
    wb.save(temp_path)
    return temp_path

def transformar_planilha_corrigida(caminho_corrigido: Path, codigo_obra: str) -> pd.DataFrame:
    """
    - Lê 'Aba1'
    - Padroniza colunas
    - Reconstrói Classe2/Classe3/ClasseComp a partir de 'unnamed: 1'
    - Regra geral: mantém linhas com exatamente 1 mês com valor
    - Para o código 1030303 NÃO há agregação; apenas força:
        * 'Cto/ Pedido/ NF' = 'Documento'
        * 'Descrição' = 'Custos com Serviços PJ - Obra'
      (mantendo valores e datas normalmente por linha)
    - Converte Data (mmm/aa) para as linhas gerais
    - Insere CC (primeira coluna)
    - Retorna DataFrame (não salva aqui)
    """
    df = pd.read_excel(caminho_corrigido, sheet_name="Aba1")
    df.columns = [str(c).strip().lower().replace('\n', ' ') for c in df.columns]
    df.columns = [c.replace('  ', ' ') for c in df.columns]

    col_codigo = 'unnamed: 1'
    possiveis_desc = [c for c in df.columns if 'descri' in c]
    if not possiveis_desc:
        raise ValueError("Coluna de Descrição não encontrada.")
    col_desc = possiveis_desc[0]

    possiveis_nf = [c for c in df.columns if 'pedido' in c or 'nf' in c]
    col_nf = possiveis_nf[0] if possiveis_nf else None

    # colunas de meses no formato 'Jan/24', 'Fev/24' etc. (comprimento 6, contém '/')
    colunas_mes = [c for c in df.columns if '/' in c and len(c) == 6]

    dados = []
    classe2 = ''
    classe3 = ''
    classecomp = ''
    in_1030303 = False  # <<< ESSENCIAL: flag de contexto

    # Varre linha a linha
    for _, row in df.iterrows():
        codigo = row.get(col_codigo, None)

        # Detectou cabeçalho / muda hierarquia?
        if pd.notna(codigo):
            try:
                codigo_str = str(int(float(codigo)))
                if len(codigo_str) == 3:
                    classe2 = codigo_str
                elif len(codigo_str) == 5:
                    classe3 = codigo_str
                elif len(codigo_str) >= 7:
                    classecomp = codigo_str
            except Exception:
                # se não converter, mantemos valores anteriores
                pass

            # Atualiza flag: estamos (ou não) dentro do 1030303
            in_1030303 = (classecomp == "1030303")

        descricao = row.get(col_desc, None)

        # ---------- REGRA GERAL (inclui itens sob 1030303 sem agregação) ----------
        if classecomp:
            qtde_valores = row[colunas_mes].notna().sum()
            if qtde_valores == 1 and (pd.notna(descricao) or in_1030303):
                for mes in colunas_mes:
                    valor = row[mes]
                    if pd.notna(valor) and valor != 0:
                        dados.append({
                            'Classe2': classe2,
                            'Classe3': classe3,
                            'ClasseComp': classecomp,
                            'Descrição': 'Custos com Serviços PJ - Obra' if in_1030303 else descricao,
                            'Cto/ Pedido/ NF': 'Documento' if in_1030303 else (row[col_nf] if col_nf else None),
                            'Valor': float(valor),
                            'Data': converter_data_ptbr(mes),
                            'CC': codigo_obra
                        })

    df_final = pd.DataFrame(dados)
    if df_final.empty:
        return df_final

    # Formatar Data (apenas onde for datetime)
    if 'Data' in df_final.columns and pd.api.types.is_datetime64_any_dtype(df_final['Data']):
        df_final['Data'] = df_final['Data'].dt.strftime('%b/%y').str.title()

    # Ordem e colunas finais
    colunas_validas = ['CC', 'Classe2', 'Classe3', 'ClasseComp', 'Descrição', 'Cto/ Pedido/ NF', 'Valor', 'Data']
    df_final = df_final.reindex(columns=colunas_validas)

    return df_final


def transformar_previsto_corrigido(caminho_corrigido: Path, codigo_obra: str) -> pd.DataFrame:
    """
    - Lê 'Aba1' já corrigida
    - Reconstrói Classe2/Classe3/ClasseComp (coluna 'unnamed: 1')
    - Para cada linha com ClasseComp (7+ dígitos), captura o Verba ('Unnamed: 4')
    - Retorna DataFrame com colunas: CC, Classe2, Classe3, ClasseComp, Verba
    """
    df = pd.read_excel(caminho_corrigido, sheet_name="Aba1")
    df.columns = [str(c).strip().lower().replace('\n', ' ') for c in df.columns]
    df.columns = [c.replace('  ', ' ') for c in df.columns]

    col_codigo = 'unnamed: 1'
    col_verba  = 'unnamed: 4'   # coluna do Verba sob o cabeçalho 'Previsto'
    if col_codigo not in df.columns or col_verba not in df.columns:
        raise ValueError("Colunas esperadas ('Unnamed: 1' e/ou 'Unnamed: 4') não encontradas.")

    dados = []
    classe2 = ''
    classe3 = ''
    classecomp = ''

    for _, row in df.iterrows():
        codigo = row.get(col_codigo, None)
        if pd.notna(codigo):
            try:
                codigo_str = str(int(float(codigo)))
                if len(codigo_str) == 3:
                    classe2 = codigo_str
                elif len(codigo_str) == 5:
                    classe3 = codigo_str
                elif len(codigo_str) >= 7:
                    classecomp = codigo_str
                    verba = row.get(col_verba, None)
                    # Sempre adiciona, mesmo que o Verba esteja vazio/NaN/0
                    try:
                        verba_val = float(verba)
                    except Exception:
                        verba_val = verba  # mantém NaN/vazio se não for numérico

                    dados.append({
                        'CC': codigo_obra,
                        'Classe2': classe2,
                        'Classe3': classe3,
                        'ClasseComp': classecomp,
                        'Verba': verba_val
                    })

            except:
                pass

    df_prev = pd.DataFrame(dados)
    # Garante colunas finais e elimina “fantasmas”
    colunas_validas = ['CC', 'Classe2', 'Classe3', 'ClasseComp', 'Verba']
    df_prev = df_prev.reindex(columns=colunas_validas)
    return df_prev

def processar_consolidado_previsto():
    linhas = []
    for item in tree_previsto.get_children():
        arquivo, cc = tree_previsto.item(item, "values")
        linhas.append((arquivo, cc))
    if not linhas:
        messagebox.showerror("Erro", "Adicione pelo menos um arquivo.")
        return
    if any(not cc for _, cc in linhas):
        messagebox.showerror("Erro", "Há arquivo sem CC. Defina CC para o(s) arquivo(s) selecionado(s).")
        return
    pasta_saida = entry_saida_prev.get().strip()
    if not pasta_saida:
        messagebox.showerror("Erro", "Selecione a pasta de saída.")
        return
    btn_processar_prev.config(state=tk.DISABLED)
    t = threading.Thread(target=_worker_consolidar_previsto, args=(linhas, Path(pasta_saida),), daemon=True)
    t.start()

def _worker_consolidar_previsto(linhas, pasta_saida: Path):
    try:
        agregados = []
        for arquivo, cc in linhas:
            p = Path(arquivo)
            corr = corrigir_planilha_entrada(p)  # mesma correção do Orçado
            try:
                df_parcial = transformar_previsto_corrigido(corr, cc)
                if df_parcial is not None and not df_parcial.empty:
                    agregados.append(df_parcial)
            finally:
                try:
                    corr.unlink(missing_ok=True)
                except Exception:
                    pass

        if not agregados:
            raise ValueError("Nenhuma linha de 'Previsto' encontrada em nenhum arquivo.")

        df_final = pd.concat(agregados, ignore_index=True)
        # Ordena opcionalmente
        ordem = ['CC', 'Classe2', 'Classe3', 'ClasseComp']
        existentes = [c for c in ordem if c in df_final.columns]
        df_final = df_final.sort_values(by=existentes)

        # Garante colunas finais
        colunas_validas = ['CC', 'Classe2', 'Classe3', 'ClasseComp', 'Verba']
        df_final = df_final.reindex(columns=colunas_validas)

        out_path = pasta_saida / "RESULTADO_PREVISTO_CONSOLIDADO.xlsx"
        df_final.to_excel(out_path, index=False)
        messagebox.showinfo("Sucesso", f"Consolidado 'Previsto' gerado:\n{out_path}")
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro em 'Previsto':\n{e}")
    finally:
        btn_processar_prev.config(state=tk.NORMAL)

def selecionar_arquivos_prev():
    arquivos = filedialog.askopenfilenames(title="Selecione as planilhas Excel (Previsto)", filetypes=[("Excel", "*.xlsx")])
    if not arquivos:
        return
    for arq in arquivos:
        tree_previsto.insert("", tk.END, values=(arq, ""))

def definir_cc_para_selecionado_prev():
    sel = tree_previsto.selection()
    cc = entry_cc_prev.get().strip()
    if not sel:
        messagebox.showerror("Erro", "Selecione uma ou mais linhas na tabela (Previsto).")
        return
    if not cc:
        messagebox.showerror("Erro", "Informe um CC no campo ao lado (Previsto).")
        return
    for item in sel:
        valores = list(tree_previsto.item(item, "values"))
        valores[1] = cc
        tree_previsto.item(item, values=valores)

def selecionar_pasta_saida_prev():
    pasta = filedialog.askdirectory(title="Selecione a pasta para salvar o consolidado (Previsto)")
    if pasta:
        entry_saida_prev.delete(0, tk.END)
        entry_saida_prev.insert(0, pasta)


# ---------- GUI / Lógica de processamento ----------
def selecionar_arquivos():
    arquivos = filedialog.askopenfilenames(
        title="Selecione as planilhas Excel",
        filetypes=[("Excel", "*.xlsx")]
    )
    if not arquivos:
        return
    for arq in arquivos:
        tree.insert("", tk.END, values=(arq, ""))

def definir_cc_para_selecionado():
    sel = tree.selection()
    cc = entry_cc.get().strip()
    if not sel:
        messagebox.showerror("Erro", "Selecione uma ou mais linhas na tabela.")
        return
    if not cc:
        messagebox.showerror("Erro", "Informe um CC no campo ao lado.")
        return
    for item in sel:
        valores = list(tree.item(item, "values"))
        valores[1] = cc
        tree.item(item, values=valores)

def selecionar_pasta_saida():
    pasta = filedialog.askdirectory(title="Selecione a pasta para salvar o consolidado")
    if pasta:
        entry_saida.delete(0, tk.END)
        entry_saida.insert(0, pasta)

def processar_consolidado():
    # Coleta (arquivo, cc) da tabela
    linhas = []
    for item in tree.get_children():
        arquivo, cc = tree.item(item, "values")
        linhas.append((arquivo, cc))

    # Valida
    if not linhas:
        messagebox.showerror("Erro", "Adicione pelo menos um arquivo.")
        return
    if any(not cc for _, cc in linhas):
        messagebox.showerror("Erro", "Há arquivo sem CC. Defina CC para o(s) arquivo(s) selecionado(s).")
        return
    pasta_saida = entry_saida.get().strip()
    if not pasta_saida:
        messagebox.showerror("Erro", "Selecione a pasta de saída.")
        return

    btn_processar.config(state=tk.DISABLED)
    t = threading.Thread(target=_worker_consolidar, args=(linhas, Path(pasta_saida),), daemon=True)
    t.start()

def _worker_consolidar(linhas, pasta_saida: Path):
    try:
        agregados = []
        for arquivo, cc in linhas:
            p = Path(arquivo)
            corr = corrigir_planilha_entrada(p)
            try:
                df_parcial = transformar_planilha_corrigida(corr, cc)
                if df_parcial is not None and not df_parcial.empty:
                    agregados.append(df_parcial)
            finally:
                # Remove temporário corrigido
                try:
                    corr.unlink(missing_ok=True)
                except Exception:
                    pass

        if not agregados:
            raise ValueError("Nenhum lançamento encontrado em nenhum arquivo.")

        df_final = pd.concat(agregados, ignore_index=True)

        # REMOVE se por ventura vier um "Custos com Serviços [FD]" no resultado
        if "Descrição" in df_final.columns:
            df_final = df_final[df_final["Descrição"] != "Custos com Serviços [FD]"]

        # Ordena opcionalmente por CC, Classe2, Classe3, ClasseComp, Data
        ordem = ['CC', 'Classe2', 'Classe3', 'ClasseComp', 'Data']
        existentes = [c for c in ordem if c in df_final.columns]
        df_final = df_final.sort_values(by=existentes)

        out_path = pasta_saida / "RESULTADO_CONSOLIDADO.xlsx"
        df_final.to_excel(out_path, index=False)

        messagebox.showinfo("Sucesso", f"Consolidado gerado:\n{out_path}")
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro:\n{e}")
    finally:
        btn_processar.config(state=tk.NORMAL)

# ---------- Montagem da Interface com ABAS ----------
root = tk.Tk()
root.title("Consolidador de Planilhas - CC por Arquivo")

notebook = ttk.Notebook(root)
notebook.pack(fill="both", expand=True)

# ---- Aba 1: Orçado (principal) ----
tab_orcado = ttk.Frame(notebook)
notebook.add(tab_orcado, text="Orçado")

frm_top = tk.Frame(tab_orcado)
frm_top.pack(fill="x", padx=10, pady=(10, 5))

tk.Label(frm_top, text="Código Externo da Obra (CC):").pack(side="left")
entry_cc = tk.Entry(frm_top, width=30)
entry_cc.pack(side="left", padx=5)
tk.Button(frm_top, text="Aplicar ao(s) selecionado(s)", command=definir_cc_para_selecionado).pack(side="left", padx=5)

frm_mid = tk.Frame(tab_orcado)
frm_mid.pack(fill="both", expand=True, padx=10, pady=5)

cols = ("arquivo", "cc")
tree = ttk.Treeview(frm_mid, columns=cols, show="headings", height=10)
tree.heading("arquivo", text="Arquivo (.xlsx)")
tree.heading("cc", text="CC")
tree.column("arquivo", width=520)
tree.column("cc", width=120)

vsb = ttk.Scrollbar(frm_mid, orient="vertical", command=tree.yview)
tree.configure(yscrollcommand=vsb.set)

tree.pack(side="left", fill="both", expand=True)
vsb.pack(side="right", fill="y")

frm_bottom = tk.Frame(tab_orcado)
frm_bottom.pack(fill="x", padx=10, pady=(5, 10))

tk.Button(frm_bottom, text="Adicionar arquivos…", command=selecionar_arquivos).pack(side="left")

tk.Label(frm_bottom, text="Pasta de saída:").pack(side="left", padx=(10, 0))
entry_saida = tk.Entry(frm_bottom, width=40)
entry_saida.pack(side="left", padx=5)
tk.Button(frm_bottom, text="Selecionar…", command=selecionar_pasta_saida).pack(side="left", padx=5)

btn_processar = tk.Button(frm_bottom, text="Processar e Consolidar", command=processar_consolidado)
btn_processar.pack(side="right")

# ---- Aba 2: Previsto ----
tab_previsto = ttk.Frame(notebook)
notebook.add(tab_previsto, text="Previsto")

frm_top_prev = tk.Frame(tab_previsto)
frm_top_prev.pack(fill="x", padx=10, pady=(10, 5))

tk.Label(frm_top_prev, text="Código Externo da Obra (CC):").pack(side="left")
entry_cc_prev = tk.Entry(frm_top_prev, width=30)
entry_cc_prev.pack(side="left", padx=5)
tk.Button(frm_top_prev, text="Aplicar ao(s) selecionado(s)", command=definir_cc_para_selecionado_prev).pack(side="left", padx=5)

frm_mid_prev = tk.Frame(tab_previsto)
frm_mid_prev.pack(fill="both", expand=True, padx=10, pady=5)

cols_prev = ("arquivo", "cc")
tree_previsto = ttk.Treeview(frm_mid_prev, columns=cols_prev, show="headings", height=10)
tree_previsto.heading("arquivo", text="Arquivo (.xlsx)")
tree_previsto.heading("cc", text="CC")
tree_previsto.column("arquivo", width=520)
tree_previsto.column("cc", width=120)

vsb_prev = ttk.Scrollbar(frm_mid_prev, orient="vertical", command=tree_previsto.yview)
tree_previsto.configure(yscrollcommand=vsb_prev.set)

tree_previsto.pack(side="left", fill="both", expand=True)
vsb_prev.pack(side="right", fill="y")

frm_bottom_prev = tk.Frame(tab_previsto)
frm_bottom_prev.pack(fill="x", padx=10, pady=(5, 10))

tk.Button(frm_bottom_prev, text="Adicionar arquivos…", command=selecionar_arquivos_prev).pack(side="left")

tk.Label(frm_bottom_prev, text="Pasta de saída:").pack(side="left", padx=(10, 0))
entry_saida_prev = tk.Entry(frm_bottom_prev, width=40)
entry_saida_prev.pack(side="left", padx=5)
tk.Button(frm_bottom_prev, text="Selecionar…", command=selecionar_pasta_saida_prev).pack(side="left", padx=5)

btn_processar_prev = tk.Button(frm_bottom_prev, text="Processar e Consolidar (Previsto)", command=processar_consolidado_previsto)
btn_processar_prev.pack(side="right")

# ---- Aba 3: Ajuda ----
tab_ajuda = ttk.Frame(notebook)
notebook.add(tab_ajuda, text="Ajuda")
ajuda_txt = (
    "Como usar:\n"
    "1) Selecione a aba correta conforme necessário (planilha de Orçado ou de Previsto).\n"
    "2) Clique em 'Adicionar arquivos…' e selecione uma ou mais planilhas .xlsx.\n"
    "3) Selecione uma obra e digite o CC no campo 'Código Externo da Obra (CC)' e clique 'Aplicar ao(s) selecionado(s)'.\n"
    "4) Escolha a pasta de saída e clique em 'Processar e Consolidar'.\n"
    "\nCaso tenha problemas e precise entrar em contato com o desenvolvedor:\n"
    "eric.rosa@elco.com.br\n"
    "\nRegras do processamento:\n"
    "- Hierarquia vem da coluna B (Unnamed: 1): 3/5/7+ dígitos => Classe2/Classe3/ClasseComp.\n"
    "- Só entram linhas com exatamente 1 mês com valor.\n"
    "- Para o código 1030303, não há agregação; apenas força Descrição e Documento conforme regras.\n"
    "- Anexe os arquivos originais sem alterações (extraídos diretamente do Mega); o programa faz as correções necessárias.\n"
    "- A coluna 'CC' é inserida como primeira coluna.\n"
    " "
    "Versão V7.0.1 - Novembro, 2025"
)
tk.Message(tab_ajuda, text=ajuda_txt, width=680).pack(padx=20, pady=20)

root.mainloop()
